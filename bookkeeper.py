import openpyxl as xl

class Bookkeeper:
    def __init__(self):
        self.INS_COL    = 'E' # the safe insertion column
        self.CAT_COL    = 'B' # category column
        self.BUD_COL    = 'D' # budget/estimate column
        self.DAT_COL    = 'C' # date column
        self.bookname   = None
        self.wb         = None
        self.ws         = None
        self.vocab      = None
        self.tabs       = {'inc' : 3, 'exp' : 22} # category-start_row mappings
        self.curr_row   = self.tabs['inc']
    
    def categorize(self, transaction):
        '''
        Check which patterns appear in the transaction description, and assign
        the appropriately mapped category tuple.
        Raises BookkeeperError if the transaction could not be categorized.
        For this function to work, call 'build_vocab()' first.
        '''
        tokens = transaction.split(',')
        desc = tokens[2].lower()
        keys = list(self.vocab)
        this_category,ii = None,0
        
        while this_category == None and ii < len(keys):
            this_key = keys[ii]
            if this_key in desc:
                this_category = self.vocab[this_key]
            ii += 1
        if this_category == None: 
            raise BookkeeperError('Failed to determine category')
        
        return this_category
    
    def get_categories(self):
        '''Returns a list of category tuples already in the current tab.'''
        categories = []
        row = self.curr_row
        value = self.ws['{}{}'.format(self.CAT_COL, row)].value
        while value != None:
            categories.append(value)
            row += 1
            value = self.ws['B{}'.format(row)].value
            
        return categories
    
    def next_free(self):
        '''
        Finds the next empty row in a tab.
        Raises BookkeeperError if there are no free rows in the tab.
        '''
        row = self.curr_row
        value = self.ws['{}{}'.format(self.CAT_COL, row)].value
        while value != None:
            row += 1
            value = self.ws['{}{}'.format(self.CAT_COL, row)].value
            if self.ws['{}{}'.format(self.CAT_COL, row+1)].value == 'Total':
                raise BookkeeperError('Current tab is full')
            
        return row
    
    def set_tab(self, key):
        try:
            self.curr_row = self.tabs[key]
        except KeyError as e:
            raise BookkeeperError('No such tab')
    
    def insert(self, date, debit, credit, category):
        '''
        Selects the appropriate tab from the category tuple and finds the
        correct row to append the debit or credit values to. 
        If the category is not in the sheet, then the next availble row is 
        located and used for insertion.
        BookkeeperError will be raised if there's no availble space in the tab.
        '''
        try:
            debit = float(debit)    
            credit = float(credit)
        except ValueError as e:
            raise BookkeeperError('Debit and Credit must be reals')
        
        # try find row with the imported category
        self.curr_row = self.tabs[category[0].lower()]
        tab_categories = self.get_categories()
        insertion_cell = None
        for ii in range(len(tab_categories)): 
            if category[1].lower() in tab_categories[ii].lower():
                insertion_cell = '{}{}'.format(self.INS_COL, self.curr_row + ii)
                
        # insert the category into a new row if need be
        if insertion_cell == None:
            free_row = self.next_free()
            self.ws['{}{}'.format(self.CAT_COL, free_row)] = category[1]
            self.ws['{}{}'.format(self.BUD_COL, free_row)] = category[2]
            insertion_cell = '{}{}'.format(self.INS_COL, free_row)
            
        # choose the > of the two, insert the amount lost/gained
        amount = debit if debit > credit else credit 
        if self.ws[insertion_cell].value == None:
            self.ws[insertion_cell] = '=+{}'.format(amount)
        else:
            self.ws[insertion_cell] = (
                self.ws[insertion_cell].value + '+{}'.format(amount)
            )
            
        # overwrite date as this is the most recent transaction in the category
        self.ws['{}{}'.format(self.DAT_COL, insertion_cell[1:])] = date
    
    def build_vocab(self, vocab_file_name):
        '''
        Builds an internal dictionary of keyword-to-category mappings which can
        be used to categorize transactions.
        The vocabulary csv file should be of the format:
        type,category,budget/estimate,keyword1,keyword2,...,keyword-n
        '''
        try:
            fl = open(vocab_file_name)
            vocab_file_lines = fl.read().split('\n')[1:] # skip header row
            fl.close()
        except IOError as e:
            raise BookkeeperError('Provided file is invalid')
        
        vocab_file_lines.pop(-1)
        mappings = dict()
        for ln in vocab_file_lines:
            data = ln.split(',')
            cat = (data[0], data[1], float(data[2])) # (type, category, budget/estimate)
            for key in data[3:]:
                mappings[key.lower()] = cat
                
        self.vocab = mappings
    
    def open(self, bookname):
        try:
            self.wb = xl.load_workbook(bookname)
            self.ws = self.wb.active 
            self.bookname = bookname
        except xl.utils.exceptions.InvalidFileException as e:
            raise BookkeeperError(str(e))
    
    def close(self):
        self.wb.save(self.bookname)
        self.wb.close()

class BookkeeperError(Exception):
    pass
